import openpyxl
import os
from openpyxl.utils import get_column_letter

def update_excel(file_path):
    if not file_path.endswith('.xlsx'):
        print("Error: El archivo debe tener la extensión '.xlsx'.")
        return
    
    if not os.path.exists(file_path):
        print("El archivo no existe. Creando uno nuevo...")
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet.append(["Nombre", "Nota", "Asignatura"])
        try:
            wb.save(file_path)
            print(f"Archivo {file_path} creado y encabezado agregado.")
        except PermissionError:
            print("Error: No se puede guardar el archivo. Asegúrate de que esté cerrado.")
            return
    else:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    
    new_row = ["Juan", "80", "Matemáticas"]
    sheet.append(new_row)
    print("Datos agregados correctamente.")
    
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                print(f"Error al procesar la celda: {e}")
        adjusted_width = max_length + 2  
        sheet.column_dimensions[column].width = adjusted_width
    
    try:
        wb.save(file_path)
        print(f"Archivo {file_path} guardado correctamente.")
    except PermissionError:
        print("Error: No se puede guardar el archivo. Asegúrate de que esté cerrado.")
    
    wb.close()

file_path = "notas.xlsx"
update_excel(file_path)

